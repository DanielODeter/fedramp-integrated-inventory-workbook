# (c) 2019 Amazon Web Services, Inc. or its affiliates. All Rights Reserved.
# License:
# This sample code is made available under the MIT-0 license. See the LICENSE file.
from datetime import datetime
import logging
import tempfile
import os, os.path
from typing import List
import boto3
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from inventory.mappers import InventoryData

_logger = logging.getLogger("inventory.reports")
_logger.setLevel(getattr(logging, os.environ.get("LOG_LEVEL", "INFO"), logging.INFO))
_current_dir_name = os.path.dirname(__file__)
_workbook_template_file_name = os.path.join(_current_dir_name, "SSP-A13-FedRAMP-Integrated-Inventory-Workbook-Template.xlsx")
_workbook_output_file_path = os.path.join(tempfile.gettempdir(), "SSP-A13-FedRAMP-Integrated-Inventory.xlsx")
DEFAULT_REPORT_WORKSHEET_FIRST_WRITEABLE_ROW_NUMBER = 3

# FedRAMP template column mappings
COL_UNIQUE_ID = 1
COL_IP_ADDRESS = 2
COL_IS_VIRTUAL = 3
COL_IS_PUBLIC = 4
COL_DNS_NAME = 5
COL_MAC_ADDRESS = 7
COL_AUTHENTICATED_SCAN = 8
COL_BASELINE_CONFIG = 9
COL_ASSET_TYPE = 12
COL_HARDWARE_MODEL = 13
COL_SOFTWARE_VENDOR = 15
COL_SOFTWARE_PRODUCT = 16
COL_IIR_DIAGRAM_LABEL = 18
COL_NETWORK_ID = 21
COL_OWNER = 22

class CreateReportCommandHandler():
    def _write_cell_if_value_provided(self, worksheet: Worksheet, column:int, row: int, value: str):
        if value is not None:
            worksheet.cell(column=column, row=row, value=value)

    def execute(self, inventory: List[InventoryData]) -> str:
        try:
            workbook = load_workbook(_workbook_template_file_name)
        except FileNotFoundError:
            _logger.error(f"Template file not found: {_workbook_template_file_name}")
            raise
        except Exception as e:
            _logger.error(f"Failed to load workbook template: {e}", exc_info=True)
            raise
        
        report_worksheet_name = os.environ.get("REPORT_WORKSHEET_NAME", "Inventory")
        
        if report_worksheet_name not in workbook.sheetnames:
            raise ValueError(f"Worksheet '{report_worksheet_name}' not found in template")
        
        report_worksheet = workbook[report_worksheet_name]
        
        try:
            rowNumber: int = int(os.environ.get("REPORT_WORKSHEET_FIRST_WRITEABLE_ROW_NUMBER", DEFAULT_REPORT_WORKSHEET_FIRST_WRITEABLE_ROW_NUMBER))
        except ValueError as e:
            _logger.error(f"Invalid row number in environment variable: {e}")
            raise ValueError("REPORT_WORKSHEET_FIRST_WRITEABLE_ROW_NUMBER must be a valid integer")

        _logger.info(f"writing {len(inventory)} rows into worksheet {report_worksheet_name} starting at row {rowNumber}")

        for inventory_row in inventory:
            self._write_cell_if_value_provided(report_worksheet, COL_UNIQUE_ID, rowNumber, inventory_row.unique_id)
            self._write_cell_if_value_provided(report_worksheet, COL_IP_ADDRESS, rowNumber, inventory_row.ip_address)
            self._write_cell_if_value_provided(report_worksheet, COL_IS_VIRTUAL, rowNumber, inventory_row.is_virtual)
            self._write_cell_if_value_provided(report_worksheet, COL_IS_PUBLIC, rowNumber, inventory_row.is_public)
            self._write_cell_if_value_provided(report_worksheet, COL_DNS_NAME, rowNumber, inventory_row.dns_name)
            self._write_cell_if_value_provided(report_worksheet, COL_MAC_ADDRESS, rowNumber, inventory_row.mac_address)
            self._write_cell_if_value_provided(report_worksheet, COL_AUTHENTICATED_SCAN, rowNumber, inventory_row.authenticated_scan_planned)
            self._write_cell_if_value_provided(report_worksheet, COL_BASELINE_CONFIG, rowNumber, inventory_row.baseline_config)
            self._write_cell_if_value_provided(report_worksheet, COL_ASSET_TYPE, rowNumber, inventory_row.asset_type)
            self._write_cell_if_value_provided(report_worksheet, COL_HARDWARE_MODEL, rowNumber, inventory_row.hardware_model)
            self._write_cell_if_value_provided(report_worksheet, COL_SOFTWARE_VENDOR, rowNumber, inventory_row.software_vendor)
            self._write_cell_if_value_provided(report_worksheet, COL_SOFTWARE_PRODUCT, rowNumber, inventory_row.software_product_name)
            self._write_cell_if_value_provided(report_worksheet, COL_IIR_DIAGRAM_LABEL, rowNumber, inventory_row.iir_diagram_label)
            self._write_cell_if_value_provided(report_worksheet, COL_NETWORK_ID, rowNumber, inventory_row.network_id)
            self._write_cell_if_value_provided(report_worksheet, COL_OWNER, rowNumber, inventory_row.owner)

            rowNumber += 1

        workbook.save(_workbook_output_file_path)

        _logger.info(f"completed saving inventory into {_workbook_output_file_path}")

        return _workbook_output_file_path

class DeliverReportCommandHandler():
    def __init__(self, s3_client=boto3.client('s3')):
        self._s3_client = s3_client

    def execute(self, report_file_name: str) -> str:
        target_path = os.environ.get("REPORT_TARGET_BUCKET_PATH")
        target_bucket = os.environ.get("REPORT_TARGET_BUCKET_NAME")
        
        if not target_path or not target_bucket:
            raise ValueError("REPORT_TARGET_BUCKET_PATH and REPORT_TARGET_BUCKET_NAME environment variables are required")
        
        # Validate target_path to prevent path traversal
        if '..' in target_path or target_path.startswith('/'):
            raise ValueError(f"Invalid target path format: {target_path}")
        
        # Validate file path to prevent path traversal
        if os.path.abspath(report_file_name) != os.path.abspath(_workbook_output_file_path):
            raise ValueError(f"Invalid report file path: {report_file_name}")
        
        # Use the validated expected path for file operations
        validated_path = _workbook_output_file_path
        
        report_filename = os.path.basename(validated_path)
        report_stem = os.path.splitext(report_filename)[0]
        report_s3_key = f"{target_path}/{report_stem}-{datetime.now().strftime('%Y-%m-%d-%H-%M-%S')}.xlsx"
        
        _logger.info(f"uploading file '{validated_path}' to bucket '{target_bucket}' with key '{report_s3_key}'")

        with open(validated_path, "rb") as object_data:
            self._s3_client.put_object(Bucket=target_bucket, Key=report_s3_key, Body=object_data)

        _logger.info(f"completed file upload")

        return f"https://{target_bucket}.s3.amazonaws.com/{report_s3_key}"